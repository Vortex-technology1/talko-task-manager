"""
Імпорт лідів та контактів VINZART в Talko CRM
Запуск: python3 import_vinzart.py
Потрібно: pip install firebase-admin openpyxl
"""

import firebase_admin
from firebase_admin import credentials, firestore
import openpyxl
import uuid
import re
from html.parser import HTMLParser
from datetime import datetime, timezone

# ── Конфіг ────────────────────────────────────────────────
COMPANY_ID = 'vinzart_mn8pnb4w'
LEADS_FILE = 'ЛідНагода (crm.lead).xlsx'
CONTACTS_FILE = 'Контакт (res.partner).xlsx'

# Service Account JSON — скачай з Firebase Console:
# Project Settings → Service Accounts → Generate new private key
SERVICE_ACCOUNT = 'serviceAccount.json'

# ── HTML stripper ──────────────────────────────────────────
class MLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.fed = []
    def handle_data(self, d):
        self.fed.append(d)
    def get_data(self):
        return ' '.join(self.fed).strip()

def strip_html(html):
    if not html: return ''
    s = MLStripper()
    s.feed(str(html))
    return s.get_data()

# ── Маппінг ────────────────────────────────────────────────
STAGE_MAP = {
    'Новий ЛІД':              'new',
    'Лід кваліфіковано':      'qualified',
    'Консультація / Зустріч': 'consultation',
    'Заміри / ТЗ отримано':   'measurements',
    'Проект / Розрахунок':    'project',
    'Пропозиція на розгляді': 'proposal',
    'Не визначився':          'thinking',
    'Виробництво':            'production',
}

SOURCE_MAP = {
    'Instagram': 'instagram',
    'Facebook':  'facebook',
    'Сарафан':   'referral',
}

# ── Ініціалізація Firebase ─────────────────────────────────
cred = credentials.Certificate(SERVICE_ACCOUNT)
firebase_admin.initialize_app(cred)
db = firestore.client()

now = datetime.now(timezone.utc)
company_ref = db.collection('companies').document(COMPANY_ID)

# ══════════════════════════════════════════════════════════
# 1. ІМПОРТ ЛІДІВ → crm_deals
# ══════════════════════════════════════════════════════════
print('=' * 50)
print('ІМПОРТ ЛІДІВ...')
wb = openpyxl.load_workbook(LEADS_FILE)
ws = wb.active

imported = 0
skipped = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    title      = str(row[0] or '').strip()
    clientName = str(row[1] or '').strip()
    email      = str(row[2] or '').strip()
    phone      = str(row[3] or '').strip()
    manager    = str(row[4] or '').strip()
    amount     = float(row[6] or 0)
    stage_lbs  = str(row[8] or 'Новий ЛІД').strip()
    source_lbs = str(row[11] or '').strip()
    note_raw   = row[14] or ''
    stage_name = str(row[15] or stage_lbs).strip()
    comment    = row[21] or ''
    tags_raw   = str(row[22] or '').strip()

    if not phone and not clientName and not title:
        skipped += 1
        continue

    note      = strip_html(note_raw)
    full_note = strip_html(comment)
    stage     = STAGE_MAP.get(stage_lbs, 'new')
    source    = SOURCE_MAP.get(source_lbs, 'manual')
    tags      = [t.strip() for t in tags_raw.split(',') if t.strip() and tags_raw != 'False'] if tags_raw else []

    deal_id = 'imp_' + uuid.uuid4().hex[:16]
    deal_title = title or clientName or 'Без назви'

    deal = {
        'id':          deal_id,
        'title':       deal_title,
        'clientName':  clientName or title,
        'email':       email,
        'phone':       phone,
        'manager':     manager,
        'amount':      amount,
        'stage':       stage,
        'stageName':   stage_name,
        'source':      source,
        'note':        note,
        'description': full_note,
        'tags':        tags,
        'isHot':       False,
        'createdAt':   now,
        'updatedAt':   now,
        'importedFrom': 'LBS Cloud',
    }

    company_ref.collection('crm_deals').document(deal_id).set(deal)
    imported += 1
    print(f'  ✓ [{imported}] {deal_title} — {phone} ({stage_name})')

print(f'\nЛіди: {imported} імпортовано, {skipped} пропущено')

# ══════════════════════════════════════════════════════════
# 2. ІМПОРТ КОНТАКТІВ → crm_clients
# ══════════════════════════════════════════════════════════
print()
print('=' * 50)
print('ІМПОРТ КОНТАКТІВ...')
wb2 = openpyxl.load_workbook(CONTACTS_FILE)
ws2 = wb2.active

imported2 = 0
skipped2 = 0

for row in ws2.iter_rows(min_row=2, values_only=True):
    name  = str(row[0] or '').strip()
    phone = str(row[1] or '').strip()

    if not phone and not name:
        skipped2 += 1
        continue

    # Якщо name містить телефон на початку — чистимо
    clean_name = re.sub(r'^\+?\d[\d\s\-]{7,}\s*', '', name).strip()
    if not clean_name:
        clean_name = name

    client_id = 'imp_' + uuid.uuid4().hex[:16]
    client = {
        'id':           client_id,
        'name':         clean_name or phone,
        'phone':        phone,
        'createdAt':    now,
        'updatedAt':    now,
        'importedFrom': 'LBS Cloud',
    }

    company_ref.collection('crm_clients').document(client_id).set(client)
    imported2 += 1
    print(f'  ✓ [{imported2}] {clean_name} — {phone}')

print(f'\nКонтакти: {imported2} імпортовано, {skipped2} пропущено')
print()
print('✅ ІМПОРТ ЗАВЕРШЕНО')
print(f'   Лідів: {imported}')
print(f'   Контактів: {imported2}')
